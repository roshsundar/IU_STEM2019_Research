# Purpose: This program will highlight the cells of patient excel file which
# are potential cases (or cases of interest) from a text file. 
# format is Sheet#(S2) + Column/patient(E) + Row(4) : S2A4-7777-00137_S2A4
# Author: Roshan Sundar

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors


wbpath = "/N/u/rosundar/Carbonate/Downloads/KEY.xlsx"
txtpath = "/N/u/rosundar/Carbonate/s2first_potential_cases.txt"

lineList = [line.rstrip('\n') for line in open(txtpath)]
cellList = [line.split('-')[0][2:] for line in lineList]

wb = load_workbook(wbpath)
ws = wb[lineList[0][:2]]

color = PatternFill(start_color='ffcc99',
                   end_color='ffcc99',
                   fill_type='solid')

for cellname in cellList:
    ws[cellname].fill = color

wb.save(wbpath)
