# Purpose: This program will go into date folders & remove non-Primary/Recurrence cases
# Author: Roshan Sundar

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string
import os
import shutil

Sheet = "Sheet_2"
sheet = "S2"
home_path = "/N/dc2/scratch/rosundar/STEM2019/DICOM/MRI/"+Sheet+"/"

wbpath = "/N/u/rosundar/Carbonate/Downloads/KEY.xlsx"
wb = load_workbook(wbpath)
ws = wb[sheet]

dates = os.listdir(home_path)
for date in dates:
    if date[:1] != 'P':
        date_path = home_path+date+"/"
        subfolders = os.listdir(date_path)
        for subfolder in subfolders:
            cell = subfolder.split('-')[0][2:]
            columnFirstCell = coordinate_from_string(cell)[0] + "1"

            if ws[columnFirstCell] == "X" or ws[cell].fill.start_color.rgb == "00000000" or (sheet not in subfolder.split('-')[0]):
                shutil.rmtree(home_path+date+"/"+subfolder)
            subfolders = os.listdir(date_path)
