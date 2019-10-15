# Purpose: This program will get approved Patient names + Primary & Recurrence dates from patient excel file
# Author: Roshan Sundar

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import datetime

wbpath = "/N/u/rosundar/Carbonate/Downloads/KEY.xlsx"

wb = load_workbook(wbpath)
ws = wb['S2'] # Change to whatever excel sheet

datesList = []
patientList = []
for c in range(1, ws.max_column):
    if (ws.cell(row = 1, column = c).value) != 'X':
        patientList.append("Patient_" + get_column_letter(c))
        for r in range(2, ws.max_row):
            if (ws.cell(row = r, column = c).fill.start_color.rgb) != '00000000':
                date = (ws.cell(row = r, column = c).value).strftime("%Y%m%d")
                datesList.append(date)

for date in datesList:
    print date

for patient in patientList:
    print patient
